#!C:\Users\Skhaju\AppData\Local\Programs\Python\Python312\python.exe
print()
import cgi
import datetime
import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import mysql.connector
import webbrowser

# Load pre-trained face detection model
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Function to detect faces in an image
def detect_faces(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)
    return faces, gray

# Function to save details to Excel with resized images embedded
def save_to_excel(details, image_size=(300, 300), folder_path="detected_faces"):
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Time', 'Image'])
    
    for timestamp, img in details:
        date_str, time_str = timestamp.split(" ")
        img_path = f"{folder_path}/{timestamp}.jpg"
        resized_img = cv2.resize(img, image_size)
        cv2.imwrite(img_path, resized_img)
        img_obj = Image(img_path)
        ws.add_image(img_obj, f'C{len(ws["C"]) + 1}')
        ws['A' + str(len(ws["A"]) + 1)] = date_str
        ws['B' + str(len(ws["B"]) + 1)] = time_str
    
    wb.save('face_recognition_details_with_images.xlsx')

# Function to save details to MySQL database (phpMyAdmin)
def save_to_mysql(details, host='localhost', user='root', password='', database='open'):
    connection = mysql.connector.connect(
        host=host,
        user=user,
        password=password,
        database=database
    )
    cursor = connection.cursor()

    for timestamp, img in details:
        date_str, time_str = timestamp.split(" ")
        img_path = f"detected_faces/{timestamp}.jpg"
        if not is_image_path_unique(cursor, img_path):
            print(f"Duplicate image path {img_path}")
            continue

        with open(img_path, "rb") as image_file:
            image_blob = image_file.read()
        query = "INSERT INTO your_table (date, time, image_path, image_blob) VALUES (%s, %s, %s, %s)"
        cursor.execute(query, (date_str, time_str, img_path, image_blob))
        connection.commit()

    connection.close()

def is_image_path_unique(cursor, img_path):
    """Returns True if the image path is unique in the table, and False otherwise."""
    query = "SELECT EXISTS(SELECT 1 FROM your_table WHERE image_path = %s)"
    cursor.execute(query, (img_path,))
    return not cursor.fetchone()[0]
# Main function
def main():
    cap = cv2.VideoCapture(0)
    details = []
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        faces, _ = detect_faces(frame)

        for (x, y, w, h) in faces:
            # Draw rectangle around the detected face
            cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)
            # Save details
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
            details.append((timestamp, frame))

        cv2.imshow('Face Recognition', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

    if details:
        save_to_excel(details, image_size=(300, 300))  # Set your desired image size here
        save_to_mysql(details)
        url = "http://localhost/opencv/index.html" 
        webbrowser.open(url)

if __name__ == "__main__":
    main()
